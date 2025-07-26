import logging
import sys
from io import BytesIO
from typing import List, Tuple, Optional, Dict, Any

import pandas as pd
from openpyxl import Workbook, load_workbook


class EnhancedExcelChunker:
    """增强的Excel分块器 - 基于ragflow实现并添加预处理和配置支持"""
    
    def __init__(self, config: Optional[Dict[str, Any]] = None):
        """
        初始化Excel分块器
        
        Args:
            config: 配置字典，支持以下选项：
                - html_chunk_rows: HTML分块行数，None表示智能计算，默认12
                - preprocess_merged_cells: 是否预处理合并单元格，默认True
                - number_formatting: 是否格式化大数字，默认True
        """
        self.config = config or {}
        self.default_chunk_rows = self.config.get('html_chunk_rows', 12)
        self.preprocess_merged = self.config.get('preprocess_merged_cells', True)
        self.number_formatting = self.config.get('number_formatting', True)
    
    @staticmethod
    def preprocess_merged_cells(file_like_object) -> BytesIO:
        """
        预处理合并单元格 - 展开合并区域并填充值
        
        Args:
            file_like_object: 文件对象或字节流
            
        Returns:
            处理后的字节流
        """
        if isinstance(file_like_object, bytes):
            file_like_object = BytesIO(file_like_object)
        
        try:
            wb = load_workbook(file_like_object)
            
            for ws in wb.worksheets:
                # 获取所有合并单元格区域（复制列表以避免迭代中修改）
                merged_ranges = list(ws.merged_cells.ranges)
                
                for rng in merged_ranges:
                    # 获取左上角单元格的值
                    top_left_value = ws.cell(rng.min_row, rng.min_col).value
                    
                    # 先解除合并
                    ws.unmerge_cells(str(rng))
                    
                    # 填充整个区域
                    for r in range(rng.min_row, rng.max_row + 1):
                        for c in range(rng.min_col, rng.max_col + 1):
                            ws.cell(r, c).value = top_left_value
            
            # 保存到新的字节流
            output = BytesIO()
            wb.save(output)
            output.seek(0)
            return output
            
        except Exception as e:
            logging.warning(f"合并单元格预处理失败: {e}，使用原始文件")
            if hasattr(file_like_object, 'seek'):
                file_like_object.seek(0)
            return file_like_object
    
    def _load_excel_to_workbook(self, file_like_object):
        """
        加载Excel文件到工作簿对象 - 改进自ragflow实现
        
        Args:
            file_like_object: 文件对象、路径或字节流
            
        Returns:
            openpyxl.Workbook对象
        """
        if isinstance(file_like_object, bytes):
            file_like_object = BytesIO(file_like_object)
        
        # 预处理合并单元格
        if self.preprocess_merged:
            file_like_object = self.preprocess_merged_cells(file_like_object)
        
        # 检测文件类型
        if hasattr(file_like_object, 'seek'):
            file_like_object.seek(0)
            file_head = file_like_object.read(4)
            file_like_object.seek(0)
            
            # 检查是否为Excel文件格式
            if not (file_head.startswith(b'PK\x03\x04') or file_head.startswith(b'\xD0\xCF\x11\xE0')):
                logging.info("检测到非Excel格式，尝试作为CSV处理")
                try:
                    df = pd.read_csv(file_like_object)
                    return self._dataframe_to_workbook(df)
                except Exception as e_csv:
                    raise Exception(f"CSV转换失败: {e_csv}")
        
        # 尝试使用openpyxl加载
        try:
            return load_workbook(file_like_object, data_only=True)
        except Exception as e:
            logging.info(f"openpyxl加载失败: {e}，尝试pandas")
            try:
                if hasattr(file_like_object, 'seek'):
                    file_like_object.seek(0)
                df = pd.read_excel(file_like_object)
                return self._dataframe_to_workbook(df)
            except Exception as e_pandas:
                raise Exception(f"pandas加载失败: {e_pandas}，原始错误: {e}")
    
    @staticmethod
    def _dataframe_to_workbook(df: pd.DataFrame) -> Workbook:
        """将DataFrame转换为Workbook对象"""
        wb = Workbook()
        ws = wb.active
        ws.title = "Data"
        
        # 写入列名
        for col_num, column_name in enumerate(df.columns, 1):
            ws.cell(row=1, column=col_num, value=column_name)
        
        # 写入数据
        for row_num, row in enumerate(df.values, 2):
            for col_num, value in enumerate(row, 1):
                ws.cell(row=row_num, column=col_num, value=value)
        
        return wb
    
    def _calculate_smart_chunk_size(self, rows: List, default_size: int = 12) -> int:
        """
        智能计算分块大小
        
        Args:
            rows: 工作表行列表
            default_size: 默认分块大小
            
        Returns:
            计算出的分块大小
        """
        if not rows or len(rows) <= 1:
            return default_size
        
        # 分析表格复杂度
        col_count = len([c for c in rows[0] if c.value])
        total_rows = len(rows) - 1  # 除去表头
        
        if col_count <= 3:  # 简单表格
            chunk_size = min(20, max(8, total_rows // 3))
        elif col_count <= 8:  # 中等复杂度
            chunk_size = min(15, max(6, total_rows // 4))
        else:  # 复杂表格
            chunk_size = min(12, max(4, total_rows // 5))
        
        return chunk_size
    
    def html_chunking(self, file_input, chunk_rows: Optional[int] = None) -> List[str]:
        """
        HTML格式分块 - 改进自ragflow实现
        
        Args:
            file_input: 文件路径、字节流或BytesIO对象
            chunk_rows: 分块行数，None表示智能计算
            
        Returns:
            HTML格式的分块列表
        """
        if isinstance(file_input, str):
            # 文件路径
            wb = self._load_excel_to_workbook(file_input)
        else:
            # 字节流
            wb = self._load_excel_to_workbook(file_input)
        
        tb_chunks = []
        
        for sheetname in wb.sheetnames:
            ws = wb[sheetname]
            rows = list(ws.rows)
            if not rows:
                continue
            
            # 确定分块大小
            if chunk_rows is None:
                if self.default_chunk_rows is None:
                    # 智能计算
                    effective_chunk_rows = self._calculate_smart_chunk_size(rows)
                else:
                    # 使用配置的默认值
                    effective_chunk_rows = self.default_chunk_rows
            else:
                # 使用指定值
                effective_chunk_rows = chunk_rows
            
            # 构建表头
            tb_rows_0 = "<tr>"
            for t in list(rows[0]):
                header_value = t.value if t.value is not None else ""
                tb_rows_0 += f"<th>{header_value}</th>"
            tb_rows_0 += "</tr>"
            
            # 分块处理
            data_rows = rows[1:]  # 排除表头
            if not data_rows:
                continue
                
            num_chunks = (len(data_rows) - 1) // effective_chunk_rows + 1
            
            for chunk_i in range(num_chunks):
                start_idx = chunk_i * effective_chunk_rows
                end_idx = min(start_idx + effective_chunk_rows, len(data_rows))
                current_chunk_rows = data_rows[start_idx:end_idx]

                # 检查当前块的行是否有内容，如果没有则跳过
                if not any(c.value is not None and str(c.value).strip() != '' for r in current_chunk_rows for c in r):
                    continue
                
                # 构建HTML表格
                tb = f"<table><caption>{sheetname}</caption>"
                tb += tb_rows_0
                
                # 添加数据行
                for r in current_chunk_rows:
                    tb += "<tr>"
                    for c in r:
                        if c.value is None:
                            tb += "<td></td>"
                        else:
                            cell_value = c.value
                            # 数字格式化
                            if (self.number_formatting and 
                                isinstance(cell_value, (int, float)) and 
                                abs(cell_value) >= 1000):
                                cell_value = f"{cell_value:,}"
                            tb += f"<td>{cell_value}</td>"
                    tb += "</tr>"
                
                tb += "</table>\n"
                tb_chunks.append(tb)
        
        return tb_chunks
    
    def row_chunking(self, file_input) -> List[str]:
        """
        行格式分块
        
        Args:
            file_input: 文件路径、字节流或BytesIO对象
            
        Returns:
            行格式的分块列表
        """
        if isinstance(file_input, str):
            wb = self._load_excel_to_workbook(file_input)
        else:
            wb = self._load_excel_to_workbook(file_input)

        row_chunks = []
        for sheetname in wb.sheetnames:
            ws = wb[sheetname]
            rows = list(ws.rows)
            if not rows: continue

            headers = [str(c.value) if c.value is not None else "" for c in rows[0]]
            
            for i in range(1, len(rows)):
                row_values = [c.value for c in rows[i]]
                
                # 跳过空行
                if all(v is None or str(v).strip() == "" for v in row_values):
                    continue

                row_text = ", ".join(
                    [
                        f"{headers[j]}: {row_values[j]}"
                        for j in range(len(row_values))
                        if row_values[j] is not None and str(row_values[j]).strip() != ""
                    ]
                )
                row_chunks.append(row_text)

        return row_chunks
    
    def chunk_excel(self, file_input, strategy: str = "html", **kwargs) -> List[str]:
        """
        Excel分块的统一入口
        
        Args:
            file_input: 文件路径、字节流或BytesIO对象
            strategy: 分块策略，"html" 或 "row"
            **kwargs: 其他参数
            
        Returns:
            分块结果列表
        """
        try:
            if strategy == "html":
                chunk_rows = kwargs.get('chunk_rows', None)
                return self.html_chunking(file_input, chunk_rows)
            elif strategy == "row":
                return self.row_chunking(file_input)
            else:
                raise ValueError(f"不支持的分块策略: {strategy}")
                
        except Exception as e:
            logging.error(f"Excel分块失败: {e}")
            raise
    
    @staticmethod
    def get_row_count(file_input) -> int:
        """
        获取Excel文件总行数 - 用于任务分配
        
        Args:
            file_input: 文件路径或字节流
            
        Returns:
            总行数
        """
        try:
            if isinstance(file_input, str):
                if file_input.split(".")[-1].lower().find("xls") >= 0:
                    wb = load_workbook(file_input, data_only=True)
                    total = 0
                    for sheetname in wb.sheetnames:
                        ws = wb[sheetname]
                        total += len(list(ws.rows))
                    return total
                
                elif file_input.split(".")[-1].lower() in ["csv", "txt"]:
                    with open(file_input, 'rb') as f:
                        content = f.read()
                    # 简单的编码检测和行数统计
                    try:
                        text = content.decode('utf-8')
                    except:
                        text = content.decode('gbk', errors='ignore')
                    return len(text.split('\n'))
            
            else:
                # 字节流处理
                if isinstance(file_input, bytes):
                    file_like_object = BytesIO(file_input)
                
                wb = load_workbook(file_like_object, data_only=True)
                total = 0
                for sheetname in wb.sheetnames:
                    ws = wb[sheetname]
                    total += len(list(ws.rows))
                return total
                
        except Exception as e:
            logging.warning(f"获取行数失败: {e}")
            return 0


# 便利函数
def create_excel_chunker(config: Optional[Dict[str, Any]] = None) -> EnhancedExcelChunker:
    """
    创建Excel分块器实例
    
    Args:
        config: 配置字典
        
    Returns:
        EnhancedExcelChunker实例
    """
    return EnhancedExcelChunker(config)


def chunk_excel_file(file_input, strategy: str = "html", config: Optional[Dict[str, Any]] = None, **kwargs) -> List[str]:
    """
    分块Excel文件的便利函数
    
    Args:
        file_input: 文件路径或字节流
        strategy: 分块策略
        config: 配置字典
        **kwargs: 其他参数
        
    Returns:
        分块结果列表
    """
    chunker = create_excel_chunker(config)
    return chunker.chunk_excel(file_input, strategy, **kwargs)


if __name__ == "__main__":
    # 测试代码
    def test_excel_chunker():
        # 创建配置
        config = {
            'html_chunk_rows': 10,
            'preprocess_merged_cells': True,
            'number_formatting': True
        }
        
        # 创建分块器
        chunker = create_excel_chunker(config)
        
        if len(sys.argv) > 1:
            file_path = sys.argv[1]
            
            # 测试HTML分块
            print("=== HTML分块测试 ===")
            html_chunks = chunker.chunk_excel(file_path, "html")
            for i, chunk in enumerate(html_chunks[:3]):  # 只显示前3个
                print(f"Chunk {i+1}:")
                print(chunk[:200] + "..." if len(chunk) > 200 else chunk)
                print()
            
            # 测试行分块
            print("=== 行分块测试 ===")
            row_chunks = chunker.chunk_excel(file_path, "row")
            for i, chunk in enumerate(row_chunks[:5]):  # 只显示前5个
                print(f"Row {i+1}: {chunk}")
        else:
            print("用法: python excel_chunker.py <excel_file_path>")
    
    test_excel_chunker() 